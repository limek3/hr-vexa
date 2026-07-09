import re
from collections.abc import Iterable
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.db.repositories.stats import AdminSearchExportRow, AdminUserSearchReportRow

HEADER_FILL = PatternFill("solid", fgColor="2B2142")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FILL = PatternFill("solid", fgColor="EDE7F6")
BLOCKED_FILL = PatternFill("solid", fgColor="FCE4EC")
ACTIVE_FILL = PatternFill("solid", fgColor="E8F5E9")
INACTIVE_FILL = PatternFill("solid", fgColor="FFF3E0")
AVAILABLE_FILL = PatternFill("solid", fgColor="E8F5E9")
PROBLEM_FILL = PatternFill("solid", fgColor="FFEBEE")
ILLEGAL_EXCEL_CHARS = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")
MAX_CELL_LENGTH = 32767


def _safe_text(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "да" if value else "нет"
    if isinstance(value, datetime):
        return value
    text = ILLEGAL_EXCEL_CHARS.sub("", str(value))
    if text.startswith(("=", "+", "-", "@")):
        text = f"'{text}"
    return text[:MAX_CELL_LENGTH]


def _join_lines(values: Iterable[object]) -> str:
    return "\n".join(str(_safe_text(value)) for value in values if str(value).strip())


def _status(value: bool | None) -> str:
    if value is None:
        return ""
    return "включен" if value else "выключен"


def _source_line(source) -> str:
    title = _safe_text(source.title or source.input_ref)
    input_ref = _safe_text(source.input_ref)
    return f"{title} | {source.access_status} | {input_ref}"


def _sheet_title(ws, title: str, subtitle: str = "") -> None:
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=16, color="2B2142")
    ws["A1"].fill = TITLE_FILL
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(color="666666")


def _write_table(ws, start_row: int, headers: list[str], rows: list[list[object]]) -> int:
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index, row in enumerate(rows, start=start_row + 1):
        for col_index, value in enumerate(row, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=_safe_text(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    if rows:
        ws.auto_filter.ref = (
            f"A{start_row}:"
            f"{get_column_letter(len(headers))}{start_row + len(rows)}"
        )
    ws.freeze_panes = f"A{start_row + 1}"
    return start_row + len(rows)


def _fit_columns(ws, *, max_width: int = 48) -> None:
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            parts = str(value).splitlines() or [""]
            max_length = max(max_length, max(len(part) for part in parts))
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), max_width)


def _style_status_cells(ws, headers: list[str], start_row: int, rows_count: int) -> None:
    header_to_col = {header: index for index, header in enumerate(headers, start=1)}
    for row_index in range(start_row + 1, start_row + rows_count + 1):
        blocked_col = header_to_col.get("Бот заблокирован")
        if blocked_col and ws.cell(row=row_index, column=blocked_col).value == "да":
            ws.cell(row=row_index, column=blocked_col).fill = BLOCKED_FILL

        status_col = header_to_col.get("Статус поиска")
        if status_col:
            cell = ws.cell(row=row_index, column=status_col)
            cell.fill = ACTIVE_FILL if cell.value == "включен" else INACTIVE_FILL

        source_status_col = header_to_col.get("Статус источника")
        if source_status_col:
            cell = ws.cell(row=row_index, column=source_status_col)
            cell.fill = AVAILABLE_FILL if cell.value == "available" else PROBLEM_FILL


def _user_rows(report_rows: list[AdminUserSearchReportRow]) -> list[list[object]]:
    users: dict[int, AdminUserSearchReportRow] = {}
    sources_total_by_user: dict[int, int] = {}
    sources_available_by_user: dict[int, int] = {}
    for row in report_rows:
        users.setdefault(row.user_id, row)
        sources_total_by_user[row.user_id] = (
            sources_total_by_user.get(row.user_id, 0) + row.sources_total
        )
        sources_available_by_user[row.user_id] = (
            sources_available_by_user.get(row.user_id, 0) + row.sources_available
        )
    return [
        [
            row.user_id,
            row.telegram_user_id,
            row.username,
            row.first_name,
            row.user_is_blocked,
            row.user_created_at,
            row.user_updated_at,
            row.user_searches_total,
            row.user_searches_active,
            sources_total_by_user.get(row.user_id, 0),
            sources_available_by_user.get(row.user_id, 0),
            row.user_matches_today,
            row.user_matches_total,
        ]
        for row in users.values()
    ]


def _search_metric_map(
    report_rows: list[AdminUserSearchReportRow],
) -> dict[int, AdminUserSearchReportRow]:
    return {row.search_id: row for row in report_rows if row.search_id is not None}


def _search_rows(
    report_rows: list[AdminUserSearchReportRow],
    search_rows: list[AdminSearchExportRow],
) -> list[list[object]]:
    metrics = _search_metric_map(report_rows)
    rows: list[list[object]] = []
    for row in search_rows:
        metric = metrics.get(row.search_id)
        rows.append(
            [
                row.user_id,
                row.telegram_user_id,
                row.username,
                row.first_name,
                row.user_is_blocked,
                row.search_id,
                row.search_title,
                _status(row.search_is_active),
                row.search_created_at,
                row.search_updated_at,
                _join_lines(row.keywords),
                _join_lines(row.minus_words),
                _join_lines(_source_line(source) for source in row.sources),
                metric.keywords_count if metric else len(row.keywords),
                metric.minus_words_count if metric else len(row.minus_words),
                metric.sources_total if metric else len(row.sources),
                metric.sources_available
                if metric
                else len([source for source in row.sources if source.access_status == "available"]),
                metric.search_matches_today if metric else 0,
                metric.search_matches_total if metric else 0,
                metric.search_hidden_total if metric else 0,
            ],
        )
    return rows


def _source_rows(search_rows: list[AdminSearchExportRow]) -> list[list[object]]:
    rows: list[list[object]] = []
    for search in search_rows:
        for source in search.sources:
            rows.append(
                [
                    search.user_id,
                    search.telegram_user_id,
                    search.username,
                    search.search_id,
                    search.search_title,
                    _status(search.search_is_active),
                    source.source_id,
                    source.telegram_id,
                    source.title,
                    source.input_ref,
                    source.source_type,
                    source.access_status,
                    source.is_active,
                ],
            )
    return rows


def _keyword_rows(search_rows: list[AdminSearchExportRow]) -> list[list[object]]:
    rows: list[list[object]] = []
    for search in search_rows:
        for value in search.keywords:
            rows.append(
                [
                    search.user_id,
                    search.telegram_user_id,
                    search.username,
                    search.search_id,
                    search.search_title,
                    "ключ",
                    value,
                ],
            )
        for value in search.minus_words:
            rows.append(
                [
                    search.user_id,
                    search.telegram_user_id,
                    search.username,
                    search.search_id,
                    search.search_title,
                    "минус-слово",
                    value,
                ],
            )
    return rows


def _prepare_sheet(
    ws,
    title: str,
    subtitle: str,
    headers: list[str],
    rows: list[list[object]],
) -> None:
    _sheet_title(ws, title, subtitle)
    table_start = 4
    _write_table(ws, table_start, headers, rows)
    _style_status_cells(ws, headers, table_start, len(rows))
    _fit_columns(ws)


def build_admin_users_workbook(
    report_rows: list[AdminUserSearchReportRow],
    search_rows: list[AdminSearchExportRow],
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Сводка"

    users_count = len({row.user_id for row in report_rows})
    blocked_count = len({row.user_id for row in report_rows if row.user_is_blocked})
    searches_count = len({row.search_id for row in report_rows if row.search_id is not None})
    active_searches_count = len(
        {
            row.search_id
            for row in report_rows
            if row.search_id is not None and row.search_is_active
        },
    )
    matches_today = sum(
        row.search_matches_today for row in report_rows if row.search_id is not None
    )
    matches_total = sum(
        row.search_matches_total for row in report_rows if row.search_id is not None
    )
    _prepare_sheet(
        ws,
        "Сводка Vexa",
        "Общая картина по пользователям, поискам и совпадениям.",
        ["Метрика", "Значение"],
        [
            ["Пользователей", users_count],
            ["Заблокировали бота", blocked_count],
            ["Поисков всего", searches_count],
            ["Поисков включено", active_searches_count],
            ["Совпадений сегодня", matches_today],
            ["Совпадений всего", matches_total],
            ["Сгенерировано", datetime.now()],
        ],
    )

    users_ws = wb.create_sheet("Пользователи")
    _prepare_sheet(
        users_ws,
        "Пользователи",
        "Одна строка = один пользователь.",
        [
            "User ID",
            "Telegram ID",
            "Username",
            "Имя",
            "Бот заблокирован",
            "Создан",
            "Обновлён",
            "Поисков всего",
            "Поисков включено",
            "Источников всего",
            "Источников доступно",
            "Совпадений сегодня",
            "Совпадений всего",
        ],
        _user_rows(report_rows),
    )

    searches_ws = wb.create_sheet("Поиски")
    _prepare_sheet(
        searches_ws,
        "Поиски",
        "Одна строка = один поиск. Здесь видно настройки пользователя.",
        [
            "User ID",
            "Telegram ID",
            "Username",
            "Имя",
            "Бот заблокирован",
            "Search ID",
            "Название поиска",
            "Статус поиска",
            "Создан",
            "Обновлён",
            "Ключевые слова",
            "Минус-слова",
            "Источники",
            "Ключей",
            "Минус-слов",
            "Источников",
            "Источников доступно",
            "Совпадений сегодня",
            "Совпадений всего",
            "Скрыто",
        ],
        _search_rows(report_rows, search_rows),
    )

    sources_ws = wb.create_sheet("Источники")
    _prepare_sheet(
        sources_ws,
        "Источники",
        "Одна строка = один источник внутри поиска.",
        [
            "User ID",
            "Telegram ID",
            "Username",
            "Search ID",
            "Название поиска",
            "Статус поиска",
            "Source ID",
            "Telegram Source ID",
            "Название источника",
            "Ссылка / username",
            "Тип",
            "Статус источника",
            "Связь активна",
        ],
        _source_rows(search_rows),
    )

    keywords_ws = wb.create_sheet("Ключи")
    _prepare_sheet(
        keywords_ws,
        "Ключи и минус-слова",
        "Одна строка = один ключ или одно минус-слово.",
        [
            "User ID",
            "Telegram ID",
            "Username",
            "Search ID",
            "Название поиска",
            "Тип",
            "Значение",
        ],
        _keyword_rows(search_rows),
    )

    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, datetime):
                    cell.number_format = "yyyy-mm-dd hh:mm"
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
